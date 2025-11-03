#!/usr/bin/env python3
"""
Fish Market Data Web Scraper
Scrapes real-time fish prices from Sri Lankan sources:
1. fisheries.gov.lk (Excel files)
2. Ceylon Fisheries Corporation
3. malupola.com

Stores data in Firebase Firestore for frontend consumption
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime, timedelta
import json
import re
import os
import tempfile
from urllib.parse import urljoin
import logging
import firebase_admin
from firebase_admin import credentials, firestore
from typing import Dict, List, Optional
import time

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Firebase (use existing credentials if already initialized)
try:
    firebase_admin.get_app()
    logger.info("Firebase already initialized")
except ValueError:
    # Firebase not initialized, need to initialize
    cred_path = os.path.join(os.path.dirname(__file__), '../serviceAccountKey.json')
    if os.path.exists(cred_path):
        cred = credentials.Certificate(cred_path)
        firebase_admin.initialize_app(cred)
        logger.info("Firebase initialized successfully")
    else:
        logger.warning("Firebase credentials not found, running in test mode")

db = firestore.client()

# Species name mapping (Sinhala/English to standardized names)
SPECIES_MAPPING = {
    'balaya': 'Skipjack Tuna',
    'skipjack': 'Skipjack Tuna',
    'kelawalla': 'Yellowfin Tuna',
    'yellowfin': 'Yellowfin Tuna',
    'yellow fin': 'Yellowfin Tuna',
    'salaya': 'Sardinella',
    'sardinella': 'Sardinella',
    'sardine': 'Sardinella',
    'thora': 'Seer Fish',
    'seer': 'Seer Fish',
    'seerfish': 'Seer Fish',
    'talapath': 'Barracuda',
    'barracuda': 'Barracuda',
    'gurulla': 'Grouper',
    'grouper': 'Grouper',
    'isso': 'Prawns',
    'prawn': 'Prawns',
    'prawns': 'Prawns',
    'shrimp': 'Prawns',
    'kakuluwo': 'Crab',
    'crab': 'Crab',
    'dhello': 'Squid',
    'squid': 'Squid',
    'cuttlefish': 'Cuttlefish',
    'kumbalawa': 'Mackerel',
    'mackerel': 'Mackerel',
    'halmassa': 'Sprats',
    'sprats': 'Sprats',
    'tuna': 'Tuna',
    'rathambala': 'Red Snapper',
    'snapper': 'Red Snapper',
}

def normalize_species_name(raw_name: str) -> Optional[str]:
    """Normalize species name to standard format"""
    if not raw_name:
        return None

    normalized = raw_name.lower().strip()

    # Direct match
    if normalized in SPECIES_MAPPING:
        return SPECIES_MAPPING[normalized]

    # Partial match
    for key, value in SPECIES_MAPPING.items():
        if key in normalized:
            return value

    # Return capitalized original if no match
    return raw_name.title()


class FisheriesDeptScraper:
    """Scraper for fisheries.gov.lk weekly fish prices (Excel files)"""

    BASE_URL = "https://www.fisheries.gov.lk"
    PRICES_URL = f"{BASE_URL}/web/index.php/en/statistics/weekly-fish-prices"

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        })

    def scrape(self) -> List[Dict]:
        """Scrape latest weekly fish prices"""
        logger.info("Starting scrape of fisheries.gov.lk")

        try:
            # Get the main page
            response = self.session.get(self.PRICES_URL, timeout=30)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, 'html.parser')

            # Find all Excel file links
            excel_links = soup.find_all('a', href=re.compile(r'\.xlsx?$'))

            if not excel_links:
                logger.warning("No Excel files found on page")
                return []

            # Get the most recent file (usually first in the list)
            latest_link = excel_links[0]
            excel_url = urljoin(self.BASE_URL, latest_link['href'])

            logger.info(f"Downloading Excel file: {excel_url}")

            # Download Excel file
            excel_response = self.session.get(excel_url, timeout=30)
            excel_response.raise_for_status()

            # Parse Excel file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(excel_response.content)
                tmp_path = tmp_file.name

            try:
                data = self._parse_excel(tmp_path)
                logger.info(f"Extracted {len(data)} price records from Excel")
                return data
            finally:
                os.unlink(tmp_path)

        except Exception as e:
            logger.error(f"Error scraping fisheries.gov.lk: {str(e)}")
            return []

    def _parse_excel(self, file_path: str) -> List[Dict]:
        """Parse Excel file and extract fish prices"""
        data = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            # Find header row (look for the row with "Sinhala Name" and "Common Name")
            header_row = None
            for idx, row in enumerate(sheet.iter_rows(max_row=10), 1):
                row_values = [str(cell.value).lower() if cell.value else '' for cell in row]
                # Look for the specific header that has "sinhala name" and column headers
                row_text = ' '.join(row_values)
                if ('sinhala' in row_text and 'name' in row_text) or \
                   ('common' in row_text and 'name' in row_text) or \
                   ('variety' in row_text and '2024' in row_text and '2025' in row_text):
                    header_row = idx
                    break

            if not header_row:
                logger.warning("Could not find header row in Excel file")
                return []

            # Extract data rows
            for row in sheet.iter_rows(min_row=header_row + 1):
                row_data = [cell.value for cell in row]

                # Skip empty rows
                if not any(row_data):
                    continue

                # Excel structure (based on fisheries.gov.lk format):
                # Column A (index 0): Serial number
                # Column B (index 1): Sinhala name
                # Column C (index 2): Common English name
                # Column D (index 3): Last year's price (2024)
                # Column E (index 4): Last week's price (2025)
                # Column F (index 5): Current week's price (2025) <- WE WANT THIS ONE

                species = None
                price = None
                market = "Unknown"

                # Get Sinhala name from column B (index 1)
                if len(row_data) > 1 and row_data[1]:
                    species_raw = str(row_data[1]).strip()
                    # Skip header-like text
                    species_lower = species_raw.lower()
                    if species_raw and not species_raw.replace('.', '').isdigit() and \
                       not any(keyword in species_lower for keyword in ['sinhala', 'variety', 'name', 'common', 'week', 'sep', 'change']):
                        species = normalize_species_name(species_raw)

                # If no Sinhala name, try English name from column C (index 2)
                if not species and len(row_data) > 2 and row_data[2]:
                    species_raw = str(row_data[2]).strip()
                    species_lower = species_raw.lower()
                    if species_raw and not species_raw.replace('.', '').isdigit() and \
                       not any(keyword in species_lower for keyword in ['sinhala', 'variety', 'name', 'common', 'week', 'sep', 'change']):
                        species = normalize_species_name(species_raw)

                # Get current week price from column F (index 5)
                if len(row_data) > 5 and row_data[5]:
                    try:
                        price_value = row_data[5]
                        # Handle numeric values
                        if isinstance(price_value, (int, float)):
                            price = float(price_value)
                        # Handle string values
                        elif isinstance(price_value, str):
                            price_match = re.search(r'(\d+(?:\.\d+)?)', str(price_value))
                            if price_match:
                                price = float(price_match.group(1))
                    except (ValueError, TypeError):
                        pass

                # Only add if we have both species and price, and price is reasonable
                # Filter out: prices too low (<50), prices that look like years (2020-2030), suspicious species names
                if species and price and price >= 50 and not (2020 <= price <= 2030):
                    species_lower = species.lower()
                    # Skip if species contains header keywords
                    if not any(keyword in species_lower for keyword in ['variety', 'sinhala', 'common', 'week', 'change', 'sep']):
                        data.append({
                        'species': species,
                        'price': price,
                        'unit': 'LKR/kg',
                        'market': market,
                        'source': 'fisheries.gov.lk',
                        'timestamp': datetime.now().isoformat(),
                        'data_type': 'weekly_average'
                    })

        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")

        return data


class CeylonFisheriesCorpScraper:
    """Scraper for Ceylon Fisheries Corporation prices"""

    BASE_URL = "https://www.cfc.gov.lk"

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        })

    def scrape(self) -> List[Dict]:
        """Scrape CFC fish prices"""
        logger.info("Starting scrape of Ceylon Fisheries Corporation")

        data = []

        # Try different location IDs (markets)
        locations = {
            '2': 'Colombo',
            '3': 'Peliyagoda',
            '4': 'Negombo',
        }

        for location_id, market_name in locations.items():
            try:
                url = f"{self.BASE_URL}/web/index.php?option=com_fishprice&view=price&location_id={location_id}&Itemid=155&lang=en"
                response = self.session.get(url, timeout=30)

                if response.status_code != 200:
                    logger.warning(f"CFC returned status {response.status_code} for {market_name}")
                    continue

                soup = BeautifulSoup(response.content, 'html.parser')

                # Look for price tables
                tables = soup.find_all('table')

                for table in tables:
                    rows = table.find_all('tr')

                    for row in rows[1:]:  # Skip header
                        cells = row.find_all(['td', 'th'])

                        if len(cells) < 2:
                            continue

                        species_cell = cells[0].get_text(strip=True)
                        price_cell = cells[-1].get_text(strip=True)

                        # Extract price
                        price_match = re.search(r'(\d+(?:\.\d+)?)', price_cell)

                        if species_cell and price_match:
                            species = normalize_species_name(species_cell)
                            price = float(price_match.group(1))

                            data.append({
                                'species': species,
                                'price': price,
                                'unit': 'LKR/kg',
                                'market': market_name,
                                'source': 'cfc.gov.lk',
                                'timestamp': datetime.now().isoformat(),
                                'data_type': 'current'
                            })

                # Be polite, add delay between requests
                time.sleep(2)

            except Exception as e:
                logger.error(f"Error scraping CFC for {market_name}: {str(e)}")

        logger.info(f"Extracted {len(data)} price records from CFC")
        return data


class MalupolaScraper:
    """Scraper for malupola.com daily wholesale prices"""

    BASE_URL = "https://malupola.com"
    PRICES_URL = f"{BASE_URL}/daily-wholesale-fish-price/"

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        })

    def scrape(self) -> List[Dict]:
        """Scrape malupola.com daily wholesale prices"""
        logger.info("Starting scrape of malupola.com")

        try:
            response = self.session.get(self.PRICES_URL, timeout=30)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, 'html.parser')

            data = []

            # Look for price listings (structure may vary)
            # Try multiple patterns

            # Pattern 1: Tables
            tables = soup.find_all('table')
            for table in tables:
                data.extend(self._parse_price_table(table))

            # Pattern 2: Lists
            price_lists = soup.find_all('ul', class_=re.compile(r'price|fish'))
            for ul in price_lists:
                data.extend(self._parse_price_list(ul))

            # Pattern 3: Divs with price class
            price_divs = soup.find_all('div', class_=re.compile(r'price|fish|product'))
            for div in price_divs:
                price_item = self._parse_price_div(div)
                if price_item:
                    data.append(price_item)

            logger.info(f"Extracted {len(data)} price records from malupola.com")
            return data

        except Exception as e:
            logger.error(f"Error scraping malupola.com: {str(e)}")
            return []

    def _parse_price_table(self, table) -> List[Dict]:
        """Parse price table"""
        data = []
        rows = table.find_all('tr')

        for row in rows[1:]:  # Skip header
            cells = row.find_all(['td', 'th'])
            if len(cells) < 2:
                continue

            species_text = cells[0].get_text(strip=True)
            price_text = cells[-1].get_text(strip=True)

            price_match = re.search(r'(\d+(?:\.\d+)?)', price_text)

            if species_text and price_match:
                data.append({
                    'species': normalize_species_name(species_text),
                    'price': float(price_match.group(1)),
                    'unit': 'LKR/kg',
                    'market': 'Peliyagoda',
                    'source': 'malupola.com',
                    'timestamp': datetime.now().isoformat(),
                    'data_type': 'wholesale'
                })

        return data

    def _parse_price_list(self, ul) -> List[Dict]:
        """Parse price list"""
        data = []
        items = ul.find_all('li')

        for item in items:
            text = item.get_text(strip=True)

            # Try to extract species and price from text
            # Format might be: "Species Name - Rs. 1000/kg"
            parts = re.split(r'[-–:]', text)

            if len(parts) >= 2:
                species_text = parts[0].strip()
                price_text = parts[-1].strip()

                price_match = re.search(r'(\d+(?:\.\d+)?)', price_text)

                if price_match:
                    data.append({
                        'species': normalize_species_name(species_text),
                        'price': float(price_match.group(1)),
                        'unit': 'LKR/kg',
                        'market': 'Peliyagoda',
                        'source': 'malupola.com',
                        'timestamp': datetime.now().isoformat(),
                        'data_type': 'wholesale'
                    })

        return data

    def _parse_price_div(self, div) -> Optional[Dict]:
        """Parse individual price div"""
        text = div.get_text(strip=True)

        # Look for price pattern
        price_match = re.search(r'Rs\.?\s*(\d+(?:\.\d+)?)', text, re.IGNORECASE)

        if not price_match:
            return None

        # Try to extract species name (text before price)
        price_pos = price_match.start()
        species_text = text[:price_pos].strip()

        if species_text and len(species_text) > 2:
            return {
                'species': normalize_species_name(species_text),
                'price': float(price_match.group(1)),
                'unit': 'LKR/kg',
                'market': 'Peliyagoda',
                'source': 'malupola.com',
                'timestamp': datetime.now().isoformat(),
                'data_type': 'wholesale'
            }

        return None


class FishMarketScraperManager:
    """Manages all scrapers and Firestore integration"""

    def __init__(self):
        self.scrapers = [
            FisheriesDeptScraper(),
            CeylonFisheriesCorpScraper(),
            MalupolaScraper()
        ]

    def scrape_all(self) -> Dict:
        """Run all scrapers and collect data"""
        logger.info("="*60)
        logger.info("Starting fish market data scraping")
        logger.info("="*60)

        all_data = []
        results = {
            'timestamp': datetime.now().isoformat(),
            'scrapers_run': 0,
            'scrapers_success': 0,
            'total_records': 0,
            'by_source': {}
        }

        for scraper in self.scrapers:
            results['scrapers_run'] += 1
            source_name = scraper.__class__.__name__

            try:
                data = scraper.scrape()

                if data:
                    all_data.extend(data)
                    results['scrapers_success'] += 1
                    results['by_source'][source_name] = len(data)
                else:
                    results['by_source'][source_name] = 0

            except Exception as e:
                logger.error(f"Error running {source_name}: {str(e)}")
                results['by_source'][source_name] = 0

        results['total_records'] = len(all_data)

        # Save to Firestore
        if all_data:
            self.save_to_firestore(all_data)

        logger.info("="*60)
        logger.info(f"Scraping complete: {results['total_records']} total records")
        logger.info(f"Success rate: {results['scrapers_success']}/{results['scrapers_run']} scrapers")
        logger.info("="*60)

        return results

    def save_to_firestore(self, data: List[Dict]):
        """Save scraped data to Firestore"""
        logger.info(f"Saving {len(data)} records to Firestore")

        try:
            # Batch write for efficiency
            batch = db.batch()
            collection_ref = db.collection('fish_market_prices')

            for item in data:
                # Create unique document ID based on species, market, source
                doc_id = f"{item['species']}_{item['market']}_{item['source']}".replace(' ', '_').lower()
                doc_ref = collection_ref.document(doc_id)

                # Add metadata
                item['last_updated'] = firestore.SERVER_TIMESTAMP

                batch.set(doc_ref, item, merge=True)

            batch.commit()
            logger.info("Successfully saved data to Firestore")

            # Also save scraping metadata
            meta_ref = db.collection('scraper_metadata').document('latest_run')
            meta_ref.set({
                'timestamp': firestore.SERVER_TIMESTAMP,
                'records_count': len(data),
                'status': 'success'
            })

        except Exception as e:
            logger.error(f"Error saving to Firestore: {str(e)}")


def main():
    """Main entry point"""
    manager = FishMarketScraperManager()
    results = manager.scrape_all()

    print("\n" + "="*60)
    print("SCRAPING RESULTS")
    print("="*60)
    print(json.dumps(results, indent=2))
    print("="*60)


if __name__ == '__main__':
    main()
